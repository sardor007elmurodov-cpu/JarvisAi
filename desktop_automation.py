"""
JARVIS Desktop Automation Scripts
Batch processing for Excel, PDF, Images
"""
import os
from pathlib import Path
from utils import setup_logger

class DesktopAutomation:
    """
    Desktop Automation for common file operations
    Excel batch processing, PDF extraction, Image conversion
    """
    def __init__(self):
        self.logger = setup_logger("DesktopAutomation")
        
    def process_excel_batch(self, folder, operation="sum_column", column="A", sheet=0):
        """
        Process multiple Excel files in a folder
        Args:
            folder: Path to folder containing Excel files
            operation: sum_column, count_rows, extract_column
            column: Column letter (A, B, C, etc.)
            sheet: Sheet index (0 = first sheet)
        """
        try:
            import openpyxl
            results = []
            
            for file_path in Path(folder).glob("*.xlsx"):
                try:
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    ws = wb.worksheets[sheet]
                    
                    if operation == "sum_column":
                        total = 0
                        for cell in ws[column]:
                            if cell.value and isinstance(cell.value, (int, float)):
                                total += cell.value
                        results.append({"file": file_path.name, "sum": total})
                        
                    elif operation == "count_rows":
                        rows = ws.max_row
                        results.append({"file": file_path.name, "rows": rows})
                        
                    wb.close()
                except Exception as e:
                    self.logger.warning(f"Skipping {file_path.name}: {e}")
                    
            return results
        except Exception as e:
            self.logger.error(f"Excel batch error: {e}")
            return []
    
    def extract_pdf_text(self, file_path, translate=False):
        """
        Extract text from PDF
        Args:
            file_path: Path to PDF file
            translate: Auto-translate to Uzbek (requires googletrans)
        """
        try:
            import PyPDF2
            
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = ""
                
                for page in reader.pages:
                    text += page.extract_text() + "\\n"
            
            if translate:
                try:
                    from googletrans import Translator
                    translator = Translator()
                    translated = translator.translate(text, dest='uz')
                    return translated.text
                except:
                    self.logger.warning("Translation failed, returning original")
                    
            return text
        except Exception as e:
            self.logger.error(f"PDF extraction error: {e}")
            return ""
    
    def convert_images_batch(self, folder, output_format="PNG", quality=95):
        """
        Convert all images in folder to specified format
        Args:
            folder: Input folder path
            output_format: PNG, JPEG, WEBP, etc.
            quality: 1-100 for JPEG
        """
        try:
            from PIL import Image
            converted = []
            supported = ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.webp']
            
            output_folder = Path(folder) / f"converted_{output_format.lower()}"
            output_folder.mkdir(exist_ok=True)
            
            for file_path in Path(folder).iterdir():
                if file_path.suffix.lower() in supported:
                    try:
                        img = Image.open(file_path)
                        
                        # Convert to RGB if saving as JPEG
                        if output_format.upper() == "JPEG" and img.mode in ('RGBA', 'P'):
                            img = img.convert('RGB')
                        
                        output_path = output_folder / f"{file_path.stem}.{output_format.lower()}"
                        
                        if output_format.upper() == "JPEG":
                            img.save(output_path, format=output_format, quality=quality)
                        else:
                            img.save(output_path, format=output_format)
                        
                        converted.append(str(output_path))
                    except Exception as e:
                        self.logger.warning(f"Failed to convert {file_path.name}: {e}")
            
            return converted
        except Exception as e:
            self.logger.error(f"Image conversion error: {e}")
            return []
    
    def batch_rename(self, folder, pattern="file_{index}", start_index=1):
        """
        Batch rename files in folder
        Args:
            folder: Folder path
            pattern: Naming pattern (use {index} for number, {original} for original name)
            start_index: Starting number
        """
        try:
            renamed = []
            files = sorted([f for f in Path(folder).iterdir() if f.is_file()])
            
            for i, file_path in enumerate(files, start=start_index):
                ext = file_path.suffix
                new_name = pattern.format(index=i, original=file_path.stem) + ext
                new_path = file_path.parent / new_name
                
                file_path.rename(new_path)
                renamed.append({"old": file_path.name, "new": new_name})
            
            return renamed
        except Exception as e:
            self.logger.error(f"Batch rename error: {e}")
            return []

if __name__ == "__main__":
    auto = DesktopAutomation()
    print("Desktop Automation ready")
